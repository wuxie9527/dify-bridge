# -*- coding: utf-8 -*-
"""
Excel 评估报表批注写回
将审核意见添加到 Excel 文件中
"""
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelAnnotator:
    """Excel 批注写回"""

    def __init__(self, file_path: str):
        """
        初始化批注器

        Args:
            file_path: Excel 文件路径
        """
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path)

    def add_comment(self, sheet_name: str, cell: str, comment: str, author: str = "审核 AI"):
        """
        添加单元格批注

        Args:
            sheet_name: Sheet 名称
            cell: 单元格坐标，如 "C3"
            comment: 批注内容
            author: 批注作者

        Returns:
            (成功标志，错误信息)
        """
        if sheet_name not in self.wb.sheetnames:
            logger.warning(f"Sheet 不存在：{sheet_name}")
            return False, f"Sheet 不存在：{sheet_name}"

        sheet = self.wb[sheet_name]

        try:
            target_cell = sheet[cell]
        except ValueError as e:
            logger.warning(f"单元格坐标无效：{cell} - {e}")
            return False, f"单元格坐标无效：{cell}"

        # 保留原有批注
        existing_comment = target_cell.comment.text if target_cell.comment else ""
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        if existing_comment:
            new_comment = f"[{author} {timestamp}]\n{comment}\n\n---\n{existing_comment}"
        else:
            new_comment = f"[{author} {timestamp}]\n{comment}"

        target_cell.comment = Comment(new_comment, author)
        logger.info(f"添加批注到 {sheet_name}!{cell}")
        return True, None

    def highlight_cell(self, sheet_name: str, cell_range: str, color: str = "FFFF00"):
        """
        高亮问题单元格

        Args:
            sheet_name: Sheet 名称
            cell_range: 单元格范围，如 "C3" 或 "C3:C10"
            color: 填充颜色（16 进制 RGB），默认黄色
        """
        if sheet_name not in self.wb.sheetnames:
            logger.warning(f"Sheet 不存在：{sheet_name}")
            return

        sheet = self.wb[sheet_name]
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # 处理单个单元格或范围
        if ":" in cell_range:
            # 范围
            for row in sheet[cell_range]:
                for cell in row:
                    cell.fill = fill
        else:
            # 单个单元格
            sheet[cell_range].fill = fill

        logger.info(f"高亮 {sheet_name}!{cell_range}")

    def create_audit_sheet(self, issues: List[Dict[str, Any]]):
        """
        创建审核意见汇总 Sheet

        Args:
            issues: 问题列表
        """
        # 删除已有的审核意见 Sheet
        if "审核意见" in self.wb.sheetnames:
            del self.wb["审核意见"]

        sheet = self.wb.create_sheet("审核意见")

        # 表头
        headers = ["序号", "问题类别", "严重程度", "位置", "问题描述", "审核依据", "修改建议"]
        sheet.append(headers)

        # 设置表头样式
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 按严重程度排序
        severity_order = {"高": 0, "中": 1, "低": 2}
        sorted_issues = sorted(issues, key=lambda x: severity_order.get(x.get("severity", "低"), 2))

        # 添加问题列表
        for i, issue in enumerate(sorted_issues, 1):
            row = [
                i,
                issue.get("category", ""),
                issue.get("severity", ""),
                issue.get("location", ""),
                issue.get("description", ""),
                issue.get("rule_reference", ""),
                issue.get("suggestion", "")
            ]
            sheet.append(row)

            # 高亮高优先级问题
            if issue.get("severity") == "高":
                for cell in sheet[f"A{i+1}:H{i+1}"]:
                    cell[0].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 自动调整列宽
        self._adjust_column_widths(sheet)

        logger.info(f"创建审核意见汇总 Sheet：{len(issues)} 条问题")

    def _adjust_column_widths(self, sheet):
        """自动调整列宽"""
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width

    def add_watermark(self, text: str = "审核版"):
        """
        添加水印

        Args:
            text: 水印文字
        """
        # 在每个 Sheet 添加水印文字
        for sheet in self.wb.worksheets:
            sheet.sheet_view.showGridLines = False
            # 简单实现：在 A1 单元格添加文字
            if sheet["A1"].value:
                sheet["A1"].value = f"{text} - {sheet['A1'].value}"

    def save(self, output_path: str):
        """
        保存文件

        Args:
            output_path: 输出文件路径
        """
        self.wb.save(output_path)
        logger.info(f"保存文件到：{output_path}")

    def close(self):
        """关闭工作簿，释放资源"""
        if self.wb:
            self.wb.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
