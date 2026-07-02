# -*- coding: utf-8 -*-
"""
Excel 评估报表提取器 - 精简版
只提取评估审核必需的数据，减少 LLM 处理负担
"""
import openpyxl
from typing import Dict, List, Any
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelExtractor:
    """Excel 数据提取器 - 只提取审核必需数据"""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=False)

    def extract_for_audit(self) -> Dict[str, Any]:
        """
        专为评估审核提取数据（精简模式）

        只返回：
        - 表格数据（移除空行空列）
        - 关键公式（合计行公式）
        - 汇总统计

        不返回：
        - file_info.path（冗余）
        - simple_checks（LLM 自己判断）
        - 完整 formulas（只需要合计行）
        - comments（评估报表通常无备注）
        """
        result = {
            "file_info": {
                "sheet_names": self.wb.sheetnames,
                "sheet_count": len(self.wb.sheetnames),
                "extract_time": datetime.now().isoformat()
            },
            "sheets": {},
            "key_formulas": {},
            "summary": {}
        }

        for sheet_name in self.wb.sheetnames:
            logger.info(f"提取 Sheet: {sheet_name}")
            sheet = self.wb[sheet_name]

            data = self._extract_sheet_data_optimized(sheet)
            if data:
                result["sheets"][sheet_name] = data

            key_formulas = self._extract_key_formulas(sheet)
            if key_formulas:
                result["key_formulas"][sheet_name] = key_formulas

            result["summary"][sheet_name] = self._extract_summary(sheet)

        logger.info(f"精简提取完成：{len(result['sheets'])} 个 Sheet")
        return result

    def extract_all(self) -> Dict[str, Any]:
        """
        完整模式提取（兼容旧接口）

        返回所有数据包括公式、备注、简单检查
        """
        result = {
            "file_info": {
                "path": self.file_path,
                "sheet_names": self.wb.sheetnames,
                "sheet_count": len(self.wb.sheetnames),
                "extract_time": datetime.now().isoformat()
            },
            "sheets": {},
            "formulas": {},
            "comments": {},
            "simple_checks": []
        }

        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            result["sheets"][sheet_name] = self._extract_sheet_data_full(sheet)
            result["formulas"][sheet_name] = self._extract_formulas_full(sheet)
            result["comments"][sheet_name] = self._extract_comments(sheet)

        result["simple_checks"] = self._run_simple_checks()
        return result

    def _extract_sheet_data_full(self, sheet) -> List[List[Any]]:
        """完整提取表格数据（保留空行）"""
        data = []
        for row in sheet.iter_rows(values_only=True):
            row_data = [v for v in row]
            if any(v is not None for v in row_data):
                data.append(row_data)
        return data

    def _extract_formulas_full(self, sheet) -> Dict[str, Dict[str, Any]]:
        """完整提取公式（所有单元格）"""
        formulas = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    formulas[cell.coordinate] = {
                        "formula": cell.value,
                        "row": cell.row,
                        "column": cell.column_letter
                    }
        return formulas

    def _extract_comments(self, sheet) -> Dict[str, Dict[str, Any]]:
        """提取单元格备注"""
        comments = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.comment:
                    comments[cell.coordinate] = {
                        "text": cell.comment.text,
                        "author": getattr(cell.comment, 'author', 'Unknown'),
                        "row": cell.row,
                        "column": cell.column_letter
                    }
        return comments

    def _run_simple_checks(self) -> List[Dict[str, Any]]:
        """简单勾稽检查"""
        checks = []
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value in ['#REF!', '#VALUE!', '#DIV/0!', '#N/A']:
                        checks.append({
                            "type": "formula_error",
                            "location": f"{sheet_name}!{cell.coordinate}",
                            "error_type": cell.value
                        })
        return checks

    def _extract_sheet_data_optimized(self, sheet) -> List[List[Any]]:
        """提取表格数据（移除空行空列）"""
        data = []

        for row in sheet.iter_rows(values_only=True):
            if all(v is None or (isinstance(v, str) and v.strip() == '') for v in row):
                continue

            row_data = [
                v if (v is not None and not (isinstance(v, str) and v.strip() == '')) else None
                for v in row
            ]
            while row_data and row_data[-1] is None:
                row_data.pop()

            if row_data:
                data.append(row_data)

        return data

    def _extract_key_formulas(self, sheet) -> Dict[str, str]:
        """只提取关键公式（合计行/汇总行）"""
        formulas = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    if any(kw in cell.value.upper() for kw in ['SUM', 'SUBTOTAL']):
                        formulas[cell.coordinate] = cell.value
                    elif row[0].value and any(kw in str(row[0].value) for kw in ['合计', '总计', '汇总']):
                        formulas[cell.coordinate] = cell.value
        return formulas

    def _extract_summary(self, sheet) -> Dict[str, Any]:
        """提取汇总统计"""
        summary = {
            "row_count": sheet.max_row,
            "col_count": sheet.max_column,
            "has_formula": False,
            "has_total_row": False,
            "total_row_index": None
        }

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).startswith('='):
                    summary["has_formula"] = True
                    break

        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if row[0] and any(kw in str(row[0]) for kw in ['合计', '总计', '汇总']):
                summary["has_total_row"] = True
                summary["total_row_index"] = idx
                break

        return summary

    def close(self):
        if self.wb:
            self.wb.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()
